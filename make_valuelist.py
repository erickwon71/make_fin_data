# -*- coding: utf-8 -*-

import os
import time
from openpyxl import load_workbook
import pandas as pd
import requests
from bs4 import BeautifulSoup as bs
import html5lib
import openpyxl
import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials

def write_header(sheet):  
    sheet['A1'] = '회사명'
    sheet['B1'] = '현재가'
    sheet['C1'] = '현재PER'
    sheet['D1'] = 'ROE'
    sheet['E1'] = 'SRIM'
    sheet['F1'] = 'PER'
    sheet['G1'] = '매력도'
    sheet['H1'] = '적정PER'

def get_corp_df():
    f_name = 'all_codes.xlsx'    

    # engine type check
    # engine = openpyxl -> xlsx file사용할때
    # default로는 xlrd (xls file일때)
    ext = os.path.splitext(f_name)[1]
    if ext == '.xlsx':
        engin_name = 'openpyxl'
    else:
        engin_name = 'xlrd'
    
    try:
        df = pd.read_excel(f_name, names=['code', 'name'], usecols=[1, 3], engine=engin_name)
    except FileNotFoundError as e:
        print(e, f_name, " is not found.")
        print(e)
    except ValueError as e:
        print(e)
    
    #print(df)
    return df

def get_gspread_sheet():
    scope = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive',
    ]

    json_file_name = 'carbide-anvil-357914-74e1d23b716b.json'

    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
    gc = gspread.authorize(credentials)

    # Valuation_통합
    spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1rg_1MR4iE2lL4G_aZUrUm0Kb5G7K-hb3LYS68rKb15Y/edit#gid=545265508'

    # 스프레스시트 문서 가져오기 
    doc = gc.open_by_url(spreadsheet_url)

    # 시트 선택하기
    worksheet = doc.worksheet('summary')
    
    return worksheet

def read_gspread_row(ws, row_num):
    data = ws.row_values(row_num)
    print(data)
    return data
    
   
def write_gspread_cell(ws, cell_name, corp_name):    
    ret_val = True
    
    #기업명 update
    ws.update_acell(cell_name, corp_name)
    time.sleep(15)

    return ret_val

if __name__=="__main__":
    print(__name__)
    
    d = datetime.datetime.now()
    fname = 'result_%d_%d_%d_%d%d%d.xlsx' %(d.year, d.month, d.day, d.hour, d.minute, d.second)
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    # make header lines
    write_header(sheet)
    wb.save(fname)
    
    dfcodes = get_corp_df()
    # corp_name = ['삼성전자']
    # corp_id = ['005930']
    # dfcodes = pd.DataFrame({'id':corp_id, 'name':corp_name})
    
    gws = get_gspread_sheet()
    
    rown = 2
    for i in range(len(dfcodes)):
        wb = load_workbook(filename = fname)
        sheet = wb.active

        #print(i, dfcodes.iloc[i, 1], dfcodes.iloc[i, 0], '', end='')

        if write_gspread_cell(gws, 'A3', dfcodes.iloc[i, 1]) == False:
            wb.save(fname)
            continue

        row_data = read_gspread_row(gws, 3)

        for i, v in enumerate(row_data):
            sheet.cell(row=rown, column=i+1, value=v)

        rown += 1
        print('')
        wb.save(fname)
        time.sleep(1)
    