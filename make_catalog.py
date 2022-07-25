# -*- coding: utf-8 -*-

import os
import time
from openpyxl import load_workbook
import pandas as pd
import requests
from bs4 import BeautifulSoup as bs
import html5lib
import openpyxl
import datetime

def get_corp_df():
    f_name = 'all_codes.xlsx'    

    # engine type check
    # engine = openpyxl -> xlsx file사용할때
    # default로는 xlrd (xls file일때)
    ext = os.path.splitext(f_name)[1]
    if ext == '.xlsx':
        engin_name = 'openpyxl'
    else:
        engin_name = 'xlrd'
    
    try:
        df = pd.read_excel(f_name, names=['code', 'name'], usecols=[1, 3], engine=engin_name)
    except FileNotFoundError as e:
        print(e, f_name, " is not found.")
        print(e)
    except ValueError as e:
        print(e)
    
    #print(df)
    return df

def get_tables(url): 
    headers = {"user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.115 Safari/537.36"}

    resp = requests.get(url, headers=headers)
    html = bs(resp.text, "lxml")
    table = html.select("table")
    try:
        table = pd.read_html(str(table))
    except:
        return None
    
    return table

def write_header(sheet):
    sheet['G1'] = '매출액'
    sheet['J1'] = '영업이익'
    sheet['M1'] = '순이익'
    sheet['P1'] = 'eps'
    sheet['S1'] = 'bps'
    sheet['V1'] = 'dps'
    sheet['Y1'] = '기말현금'
    sheet['AB1'] = 'FCF'
    
    sheet['A2'] = '회사명'
    sheet['B2'] = 'code'
    sheet['C2'] = '제품'
    sheet['D2'] = '시총'
    sheet['E2'] = '주식수'
    sheet['F2'] = '대주주지분'
    
    sheet['G2'] = '2019'
    sheet['H2'] = '2020'
    sheet['I2'] = '2021'

    sheet['J2'] = '2019'
    sheet['K2'] = '2020'
    sheet['L2'] = '2021'
   
    sheet['M2'] = '2019'
    sheet['N2'] = '2020'
    sheet['O2'] = '2021'

    sheet['P2'] = '2019'
    sheet['Q2'] = '2020'
    sheet['R2'] = '2021'

    sheet['S2'] = '2019'
    sheet['T2'] = '2020'
    sheet['U2'] = '2021'

    sheet['V2'] = '2019'
    sheet['W2'] = '2020'
    sheet['X2'] = '2021'

    sheet['Y2'] = '2019'
    sheet['Z2'] = '2020'
    sheet['AA2'] = '2021'

    sheet['AB2'] = '2019'
    sheet['AC2'] = '2020'
    sheet['AD2'] = '2021'

    

if __name__=="__main__":
    print(__name__)
    
    d = datetime.datetime.now()
    fname = 'result_%d_%d_%d_%d%d%d.xlsx' %(d.year, d.month, d.day, d.hour, d.minute, d.second)
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    # make header lines
    write_header(sheet)
    wb.save(fname)
    
    dfcodes = get_corp_df()

    #corp_name = ['삼성전자']
    #corp_id = ['005930']
    #dfcodes = pd.DataFrame({'id':corp_id, 'name':corp_name})
    rown = 3
    for i in range(len(dfcodes)):
        wb = load_workbook(filename = fname)
        sheet = wb.active

        print(i, dfcodes.iloc[i, 1], dfcodes.iloc[i, 0], '', end='')
        sheet.cell(row=rown, column=1, value=dfcodes.iloc[i, 1]) #회사명
        sheet.cell(row=rown, column=2, value=dfcodes.iloc[i, 0]) #code
        
        corp_i = dfcodes.iloc[i, 0]

        main_url = 'https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode=A{}&cID=&MenuYn=Y&ReportGB=&NewMenuID=101&stkGb=701'.format(corp_i)
        finance_url = 'https://comp.fnguide.com/SVO2/ASP/SVD_Finance.asp?pGB=1&gicode=A{}&cID=&MenuYn=Y&ReportGB=&NewMenuID=103&stkGb=701'.format(corp_i)
        corp_url = 'https://comp.fnguide.com/SVO2/ASP/SVD_Corp.asp?pGB=1&gicode=A{}&cID=&MenuYn=Y&ReportGB=&NewMenuID=102&stkGb=701'.format(corp_i)
        invest_url = 'https://comp.fnguide.com/SVO2/ASP/SVD_Invest.asp?pGB=1&gicode=A{}&cID=&MenuYn=Y&ReportGB=&NewMenuID=105&stkGb=701'.format(corp_i)
        
        main_tbl = get_tables(main_url)
        if main_tbl == None:
            wb.save(fname)
            continue

        finance_tbl = get_tables(finance_url)
        if finance_tbl == None:
            wb.save(fname)
            continue

        corp_tbl = get_tables(corp_url)
        if corp_tbl == None:
            wb.save(fname)
            continue

        invest_tbl = get_tables(invest_url)
        if invest_tbl == None:
            wb.save(fname)
            continue

        # 주요제품     
        #print(corp_tbl[2])
        try:
            print('', corp_tbl[2].iloc[0, 0], end='')
            sheet.cell(row=rown, column=3, value=corp_tbl[2].iloc[0, 0]) #제품
        except:
            print('', 'NaN', end='')

        #시총, 발행주수
        #print(main_tbl[0]])
        try:
            print('', main_tbl[0].iloc[4, 1], main_tbl[0].iloc[6, 1].split('/')[0], end='')
            sheet.cell(row=rown, column=4, value=main_tbl[0].iloc[4, 1]) #시총
            sheet.cell(row=rown, column=5, value=main_tbl[0].iloc[6, 1].split('/')[0]) #주식수
        except:
            print('', '0', '0', end='')
        # 대주주지분
        #print(main_tbl[3])
        try:
            print('', main_tbl[3].iloc[0, 2], end='')
            sheet.cell(row=rown, column=6, value=main_tbl[3].iloc[0, 2]) #대주주지분
        except:
            print('', '0', end='')
        
        #손익
        #print(main_tbl[10])
        try:
            #매출액
            print('', main_tbl[10].iloc[0, 1], main_tbl[10].iloc[0, 2], main_tbl[10].iloc[0, 3], end='')
            sheet.cell(row=rown, column=7, value=main_tbl[10].iloc[0, 1])
            sheet.cell(row=rown, column=8, value=main_tbl[10].iloc[0, 2])
            sheet.cell(row=rown, column=9, value=main_tbl[10].iloc[0, 3])
            
            #영업이익
            print('', main_tbl[10].iloc[1, 1], main_tbl[10].iloc[1, 2], main_tbl[10].iloc[1, 3], end='')
            sheet.cell(row=rown, column=10, value=main_tbl[10].iloc[1, 1])
            sheet.cell(row=rown, column=11, value=main_tbl[10].iloc[1, 2])
            sheet.cell(row=rown, column=12, value=main_tbl[10].iloc[1, 3])
            
            #순이익
            print('', main_tbl[10].iloc[3, 1], main_tbl[10].iloc[3, 2], main_tbl[10].iloc[3, 3], end='')
            sheet.cell(row=rown, column=13, value=main_tbl[10].iloc[3, 1])
            sheet.cell(row=rown, column=14, value=main_tbl[10].iloc[3, 2])
            sheet.cell(row=rown, column=15, value=main_tbl[10].iloc[3, 3])

            #eps
            print('', main_tbl[10].iloc[18, 1], main_tbl[10].iloc[18, 2], main_tbl[10].iloc[18, 3], end='')
            sheet.cell(row=rown, column=16, value=main_tbl[10].iloc[18, 1])
            sheet.cell(row=rown, column=17, value=main_tbl[10].iloc[18, 2])
            sheet.cell(row=rown, column=18, value=main_tbl[10].iloc[18, 3])

            #bps
            print('', main_tbl[10].iloc[19, 1], main_tbl[10].iloc[19, 2], main_tbl[10].iloc[19, 3], end='')
            sheet.cell(row=rown, column=19, value=main_tbl[10].iloc[19, 1])
            sheet.cell(row=rown, column=20, value=main_tbl[10].iloc[19, 2])
            sheet.cell(row=rown, column=21, value=main_tbl[10].iloc[19, 3])
    
            #dps
            print('', main_tbl[10].iloc[20, 1], main_tbl[10].iloc[20, 2], main_tbl[10].iloc[20, 3], end='')
            sheet.cell(row=rown, column=22, value=main_tbl[10].iloc[20, 1])
            sheet.cell(row=rown, column=23, value=main_tbl[10].iloc[20, 2])
            sheet.cell(row=rown, column=24, value=main_tbl[10].iloc[20, 3])

        except:
            print('', '0', '0', '0', end='')
            print('', '0', '0', '0', end='')
            print('', '0', '0', '0', end='')
            print('', '0', '0', '0', end='')
            print('', '0', '0', '0', end='')
            print('', '0', '0', '0', end='')
            pass


        # 기말현금
        #print('')
        #print(finance_tbl[4])
        try:
            print('', finance_tbl[4].iloc[21, 1], finance_tbl[4].iloc[21, 2], finance_tbl[4].iloc[21, 3], end='')
            sheet.cell(row=rown, column=25, value=finance_tbl[4].iloc[21, 1])
            sheet.cell(row=rown, column=26, value=finance_tbl[4].iloc[21, 2])
            sheet.cell(row=rown, column=27, value=finance_tbl[4].iloc[21, 3])
        except:
            print('', '0', '0', '0', end='')
        #재무상태
        #print('')
        #print(finance_tbl[2])
        #try:
        #    print('', finance_tbl[2].iloc[0, 1], finance_tbl[2].iloc[0, 2], finance_tbl[2].iloc[0, 3], end='')
        #except:
        #    print('', '0', '0', '0', end='')

        #투자지표 FCF
        #print('')
        #print(invest_tbl[1])
        try:
            print('', invest_tbl[1].iloc[22, 2], invest_tbl[1].iloc[22, 3], invest_tbl[1].iloc[22, 4], end='')
            sheet.cell(row=rown, column=28, value=invest_tbl[1].iloc[22, 2])
            sheet.cell(row=rown, column=29, value=invest_tbl[1].iloc[22, 3])
            sheet.cell(row=rown, column=30, value=invest_tbl[1].iloc[22, 4])
        except:
            print('', '0', '0', '0', end='')
        
        rown += 1
        print('')
        wb.save(fname)
        time.sleep(1)
    